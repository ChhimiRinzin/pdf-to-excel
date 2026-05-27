import streamlit as st
import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO
import os

st.set_page_config(page_title="AWP 5.2 Automation", layout="centered", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
* { font-family: 'DM Sans', sans-serif; }

[data-testid="stAppViewContainer"] { background: #f8fafc; min-height: 100vh; }
[data-testid="stHeader"] { background: transparent; }
.block-container { max-width: 720px; padding-top: 0rem !important; padding-bottom: 2rem !important; }
#MainMenu, footer { visibility: hidden; }

/* Hero */
.hero { text-align: center; padding: 0.2rem 0 0.5rem; }
.hero .badge { display: inline-flex; align-items: center; gap: 6px; background: #dcfce7; color: #15803d; border: 1px solid #bbf7d0; border-radius: 999px; padding: 5px 14px; font-size: 12px; font-weight: 600; margin-bottom: 6px; }
.hero .badge-dot { width: 7px; height: 7px; border-radius: 50%; background: #22c55e; animation: pulse 2s infinite; }
[data-testid="stImage"] { margin: -8px 0 -6px 0 !important; }
@keyframes pulse { 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:.5;transform:scale(1.35)} }
.hero h1 { font-size: 34px; font-weight: 700; line-height: 1.18; letter-spacing: -.04em; color: #0f172a; margin: 0 0 0.5px 0; }
.hero h1 span.green { color: #16a34a; }
.hero p { font-size: 15px; line-height: 1.7; color: #64748b; margin: 0; }

/* Steps bar */
.steps-bar { display: flex; align-items: center; justify-content: center; background: #fff; border: 1px solid #e2e8f0; border-radius: 18px; padding: 12px 20px; margin: 1rem 0; box-shadow: 0 1px 2px rgba(0,0,0,0.05); }
.step-item { display: flex; align-items: center; justify-content: center; gap: 10px; flex: 1; }
.step-num { width: 32px; height: 32px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 13px; font-weight: 700; border: 2px solid #dbe4ee; background: #f8fafc; color: #64748b; }
.step-num.done   { background: #16a34a; border-color: #16a34a; color: white; }
.step-num.active { background: linear-gradient(135deg,#16a34a,#15803d); border-color: #15803d; color: white; }
.step-text { font-size: 13px; font-weight: 500; color: #64748b; }
.step-text.active { color: #0f172a; }
.step-text.done   { color: #15803d; }
.step-connector      { width: 44px; height: 2px; background: #e2e8f0; border-radius: 999px; }
.step-connector.done { background: #22c55e; }

/* Dropzone */
[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important; border: 2px dashed #cbd5e1 !important;
    border-radius: 24px !important; padding: 2rem 2rem 1.8rem !important;
    box-shadow: 0 4px 20px rgba(15,23,42,.04) !important; transition: all 0.2s ease !important;
    min-height: 280px !important; display: flex !important; flex-direction: column-reverse !important;
    align-items: center !important; justify-content: center !important; gap: 18px !important;
}
[data-testid="stFileUploaderDropzone"]:hover { border-color: #16a34a !important; background: #f0fdf4 !important; transform: translateY(-3px); }
[data-testid="stFileUploaderDropzoneInstructions"]::before { content: ""; width: 58px; height: 58px; display: block; margin-bottom: 18px; background-image: url("data:image/svg+xml,%3Csvg width='58' height='58' viewBox='0 0 58 58' fill='none' xmlns='http://www.w3.org/2000/svg'%3E%3Ccircle cx='29' cy='29' r='29' fill='%23dcfce7'/%3E%3Cpath d='M29 18v14M22 25l7-7 7 7' stroke='%2315803d' stroke-width='2.5' stroke-linecap='round' stroke-linejoin='round'/%3E%3Cpath d='M18 39c-3 0-5-2-5-5 0-2.7 2-4.8 4.6-5.5.8-3.5 3.9-6.2 7.7-6.2 4.3 0 7.7 3.4 7.7 7.7h.8c2.8 0 5 2 5 5s-2 5-5 5H18z' stroke='%2315803d' stroke-width='2.2' stroke-linecap='round' fill='none'/%3E%3C/svg%3E"); background-size: contain; background-repeat: no-repeat; background-position: center; }
[data-testid="stFileUploaderDropzoneInstructions"] div[data-testid="stMarkdownContainer"],
[data-testid="stFileUploaderDropzoneInstructions"] > div:first-child > span { display: none !important; }
[data-testid="stFileUploaderDropzoneInstructions"]::after { content: "Drag & drop your PDF here\A\ASupports PDF files up to 200MB"; white-space: pre-line; font-size: 15px; line-height: 1.8; color: #64748b; text-align: center; }
[data-testid="stFileUploaderDropzone"] button { background: linear-gradient(135deg,#16a34a,#15803d) !important; border-radius: 40px !important; font-weight: 600 !important; padding: 10px 20px !important; width: auto !important; min-width: 180px; margin-top: 16px !important; }
[data-testid="stFileUploaderDropzone"] button:hover { transform: translateY(-1px); background: linear-gradient(135deg,#22c55e,#16a34a) !important; }
[data-testid="stFileUploaderFile"] { display: none !important; }

/* Button centering wrapper */
.btn-center { display: flex; justify-content: center; margin: 0.5rem 0; }

/* All stButtons — green, fixed width */
.stButton > button {
    background: linear-gradient(135deg, #16a34a, #15803d) !important;
    border: none !important; border-radius: 40px !important;
    font-weight: 600 !important; color: white !important;
    transition: all 0.25s ease !important;
    width: auto !important; min-width: 200px; max-width: 280px;
    padding: 10px 32px !important;
    box-shadow: 0 4px 12px rgba(22,163,74,0.3);
    display: block; margin: 0 auto;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #22c55e, #16a34a) !important;
    transform: translateY(-2px);
    box-shadow: 0 10px 20px -5px rgba(22,163,74,0.5) !important;
    color: white !important;
}

/* Remove / Start Over button — muted style */
.remove-btn .stButton > button {
    background: #f1f5f9 !important;
    color: #64748b !important;
    border: 1px solid #e2e8f0 !important;
    box-shadow: none !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    min-width: 120px; max-width: 160px;
    padding: 8px 20px !important;
}
.remove-btn .stButton > button:hover {
    background: #fee2e2 !important;
    color: #dc2626 !important;
    border-color: #fecaca !important;
    box-shadow: none !important;
    transform: translateY(-1px);
}

/* Download button — green, fixed width */
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #16a34a, #15803d) !important;
    border: none !important; border-radius: 40px !important;
    font-weight: 600 !important; transition: all 0.25s ease !important;
    width: auto !important; min-width: 220px; max-width: 300px;
    padding: 10px 32px !important;
    box-shadow: 0 4px 12px rgba(22,163,74,0.3);
    display: block; margin: 0 auto;
}
[data-testid="stDownloadButton"] button:hover {
    background: linear-gradient(135deg, #22c55e, #16a34a) !important;
    transform: translateY(-2px);
    box-shadow: 0 10px 20px -5px rgba(22,163,74,0.5) !important;
}

/* Input */
[data-testid="stTextInput"] input { border-radius: 40px; border: 1px solid #e2e8f0; padding: 10px 16px; font-size: 14px; }

/* Upload success */
.upload-success { background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 20px; padding: 1rem 1.2rem; text-align: center; margin: 0.5rem 0; animation: fadeInUp 0.3s ease; }
.upload-success .filename { font-weight: 600; color: #166534; word-break: break-all; }

/* Result box */
.result-box { background: linear-gradient(135deg, #f0fdf4, #dcfce7); border-radius: 24px; padding: 1.5rem; margin-top: 1.5rem; border: 1px solid #bbf7d0; text-align: center; animation: fadeInUp 0.4s ease; }
.result-title { font-weight: 700; font-size: 18px; color: #166534; margin-bottom: 1rem; }
@keyframes fadeInUp { from{opacity:0;transform:translateY(15px)} to{opacity:1;transform:translateY(0)} }

/* Footer */
.app-footer { text-align: center; font-size: 12px; color: #94a3b8; border-top: 1px solid #e2e8f0; padding-top: 1.5rem; margin-top: 2rem; }

@media (max-width: 640px) {
    .hero h1 { font-size: 24px; }
    .steps-bar { flex-wrap: wrap; gap: 12px; }
    .step-connector { display: none; }
}
</style>
""", unsafe_allow_html=True)

# ===================================================================
# Helpers
# ===================================================================
def clean_title(t):
    return re.sub(r'[\/*?:\[\]]', '', t).strip()[:31]

def clean_cell(v):
    return "" if v is None else str(v).replace("\n", " ").strip()

def split_assertions(text):
    if not text:
        return []
    parts = re.split(r'\s+and\s+|\s*&\s*|\s*/\s*|,\s*|;\s*', text, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p.strip()]

def extract_header_info(all_tables):
    info = {k: "" for k in ["entity_name","audit_period","assessed_name","assessed_designation",
                              "assessed_date","reviewed_name","reviewed_designation","reviewed_date"]}
    for table in all_tables:
        if not table:
            continue
        for ri, row in enumerate(table):
            rt = [clean_cell(c) for c in row]
            for i, cell in enumerate(rt):
                lc = cell.lower()
                def next_val(col_offset=1):
                    vals = [x for x in rt[i+col_offset:] if x]
                    if vals: return vals[0]
                    if ri + 1 < len(table):
                        nxt = [clean_cell(c) for c in table[ri+1]]
                        v = [x for x in nxt if x]
                        if v: return v[0]
                    return ""
                if "name of the entity" in lc:
                    info["entity_name"] = next_val() or re.sub(r'name of the entity', '', cell, flags=re.IGNORECASE).strip(" :-")
                elif "period of audit" in lc:
                    info["audit_period"] = next_val() or (re.search(r'(\d{4}[-/]\d{2,4})', " ".join(rt)) or type('', (), {'group': lambda s,x: ""})()).group(1)
                elif lc in ("name:", "name"):
                    vals = [x for x in rt[i+1:] if x]
                    if len(vals) >= 1: info["assessed_name"] = vals[0]
                    if len(vals) >= 2: info["reviewed_name"] = vals[1]
                elif lc in ("designation:", "designation"):
                    vals = [x for x in rt[i+1:] if x]
                    if len(vals) >= 1: info["assessed_designation"] = vals[0]
                    if len(vals) >= 2: info["reviewed_designation"] = vals[1]
                elif lc in ("date:", "date"):
                    vals = [x for x in rt[i+1:] if x]
                    if len(vals) >= 1: info["assessed_date"] = vals[0]
                    if len(vals) >= 2: info["reviewed_date"] = vals[1]
    return info

def generate_excel(pdf_file):
    with pdfplumber.open(pdf_file) as pdf:
        all_tables = [t for page in pdf.pages for t in (page.extract_tables() or [])]

    header_info = extract_header_info(all_tables)
    relevant_tables = [t for t in all_tables if t and len(t[0]) >= 5]

    JUNK = {"material","risks","name:","date:","designation:","name of the entity",
            "prepared by","reviewed","reviewed & agreed by","period of audit","awp","not significant","1"}

    parsed_data = {}
    current_class = None

    for table in relevant_tables:
        for row in table:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            if row[0] is not None:
                raw = clean_title(clean_cell(row[0]))
                if raw and raw.lower() not in JUNK and not any(raw.lower().startswith(j) for j in JUNK):
                    current_class = raw
                    parsed_data.setdefault(current_class, [])
            if not current_class:
                continue
            assertions = split_assertions(clean_cell(row[3]) if len(row) > 3 else "") or [""]
            substantive = next((clean_cell(row[idx]) for idx in range(11, 4, -1)
                                if len(row) > idx and clean_cell(row[idx])), "")
            parsed_data[current_class].append({
                "Risk": clean_cell(row[1]) if len(row) > 1 else "",
                "Assertions": assertions,
                "Control Activity": clean_cell(row[4]) if len(row) > 4 else "",
                "Substantive Testing Procedures": substantive,
            })

    if not parsed_data:
        st.warning("⚠️ No audit tables found. The PDF may use a different format or contain scanned images.")

    wb = Workbook()
    wb.remove(wb.active)
    TH, MD = Side(style="thin"), Side(style="medium")

    def tb():
        return Border(left=TH, right=TH, top=TH, bottom=TH)

    def outer_med(ws, r1, r2, c1, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = Border(
                    left=MD if c==c1 else TH, right=MD if c==c2 else TH,
                    top=MD if r==r1 else TH, bottom=MD if r==r2 else TH,
                )

    def fill(hex_):
        return PatternFill(start_color=hex_, end_color=hex_, fill_type="solid")

    TITLE  = fill("1F4E79"); BLUE  = fill("D9EAF7"); BHDR  = fill("2E75B6")
    GREY   = fill("D9D9D9"); YELLOW= fill("FFD966"); PINK  = fill("C55A11")
    GREEN  = fill("C6EFCE"); WHITE = fill("FFFFFF"); STRIPE= fill("EBF3FB")

    def sc(cell, bold=False, f=None, ha="left", va="center", wrap=True, sz=10, color="000000"):
        cell.font      = Font(bold=bold, size=sz, color=color, name="Arial")
        cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        cell.border    = tb()
        if f: cell.fill = f

    for sheet_name, records in parsed_data.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        seen_risks = {}
        for entry in records:
            rt = entry["Risk"].strip()
            if not rt: continue
            seen_risks.setdefault(rt, [])
            for a in entry["Assertions"]:
                if a and a not in seen_risks[rt]:
                    seen_risks[rt].append(a)

        rac = [(r, a) for r, alist in seen_risks.items() for a in (alist or [""])]
        if not rac:
            rac = [("Risk", "Relevant Assertions")]

        DYN_START = 7
        REM_COL   = DYN_START + len(rac)
        LAST_COL  = REM_COL

        for col, w in [("A",2),("B",20),("C",26),("D",28),("E",32),("F",20),("G",26),("H",14),("I",14)]:
            ws.column_dimensions[col].width = w
        for ci in range(DYN_START, REM_COL):
            ws.column_dimensions[get_column_letter(ci)].width = 22
        ws.column_dimensions[get_column_letter(REM_COL)].width = 28

        ws.row_dimensions[1].height = 28
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=LAST_COL)
        ws["B1"] = f"AWP 5.2 Performing Substantive Audit Procedures - {sheet_name}"
        sc(ws["B1"], bold=True, sz=13, ha="center", f=TITLE, color="FFFFFF")

        for r, lbl, key in [(2,"Name of the Entity :","entity_name"),(3,"Period of Audit :","audit_period")]:
            ws.row_dimensions[r].height = 22
            sc(ws.cell(r, 2, lbl), bold=True, f=BLUE, wrap=False)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)
            ws.cell(r, 3, header_info[key])
            for c in range(2, LAST_COL + 1):
                ws.cell(r, c).border = tb()
                ws.cell(r, c).font   = Font(size=10, name="Arial")
                if c >= 3:
                    ws.cell(r, c).alignment = Alignment(wrap_text=False, vertical="center", horizontal="left")
            sc(ws.cell(r, 2), bold=True, f=BLUE, wrap=False)
        outer_med(ws, 2, 3, 2, LAST_COL)

        ws.row_dimensions[4].height = 6

        for r in range(5, 9):
            ws.row_dimensions[r].height = 22

        for merge, val, ref in [
            ("B5:C5","Prepared by","B5"), ("D5:E5","Signature","D5"),
            ("F5:G5","Reviewed & agreed by","F5"), ("H5:I5","Signature","H5"),
        ]:
            ws.merge_cells(merge)
            sc(ws[ref], bold=True, f=BHDR, ha="center", color="FFFFFF", sz=10)
            ws[ref] = val

        for r, lbl, ak, rk in [
            (6,"Name","assessed_name","reviewed_name"),
            (7,"Designation","assessed_designation","reviewed_designation"),
            (8,"Date","assessed_date","reviewed_date"),
        ]:
            for col, key, f_ in [(2,lbl,GREY),(6,lbl,GREY)]:
                sc(ws.cell(r, col, key), bold=True, f=f_, wrap=False)
            for col, key in [(3, ak),(7, rk)]:
                c = ws.cell(r, col, header_info[key])
                c.alignment = Alignment(wrap_text=False, vertical="center", horizontal="left")
                c.font = Font(size=10, name="Arial"); c.border = tb()

        ws.merge_cells(start_row=6, start_column=4, end_row=8, end_column=5)
        ws.merge_cells(start_row=6, start_column=8, end_row=8, end_column=9)
        for r in range(5, 9):
            for c in range(2, 10):
                ws.cell(r, c).border = tb()
        outer_med(ws, 5, 8, 2, 9)

        ws.row_dimensions[9].height = 6

        ws.row_dimensions[10].height = 20
        ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=LAST_COL)
        ws["B10"] = "STEP 1 : Trace risks, control activity, substantive audit procedures and relevant audit assertions"
        sc(ws["B10"], bold=True, sz=10, f=PINK, ha="left", color="FFFFFF")

        ws.row_dimensions[11].height = 18
        ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=LAST_COL)
        ws["B11"] = f"Significant COTABD:  {sheet_name}"
        sc(ws["B11"], bold=True, sz=10, f=BLUE, ha="left")

        ws.row_dimensions[12].height = 32
        for ci, hdr in [(2,"Risk Description"),(3,"Relevant Assertions"),(4,"Control Activity"),(5,"Substantive Testing Procedures")]:
            sc(ws.cell(12, ci, hdr), bold=True, f=BHDR, ha="center", color="FFFFFF", sz=10)
        outer_med(ws, 12, 12, 2, 5)

        s1_start = 13
        for i, entry in enumerate(records):
            rf = STRIPE if i % 2 else WHITE
            ws.row_dimensions[s1_start + i].height = 48
            for ci, val in zip([2,3,4,5],[
                entry["Risk"], " / ".join(entry["Assertions"]),
                entry["Control Activity"], entry["Substantive Testing Procedures"],
            ]):
                c = ws.cell(s1_start + i, ci, val)
                c.font = Font(size=10, name="Arial")
                c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                c.border = tb(); c.fill = rf

        s1_end = s1_start + max(len(records) - 1, 0)
        if records:
            outer_med(ws, s1_start, s1_end, 2, 5)

        s2_title = s1_end + 2
        ws.row_dimensions[s2_title - 1].height = 6
        ws.row_dimensions[s2_title].height = 20
        ws.merge_cells(start_row=s2_title, start_column=2, end_row=s2_title, end_column=LAST_COL)
        ws.cell(s2_title, 2, "STEP 2 : Substantive audit procedures performed")
        sc(ws.cell(s2_title, 2), bold=True, f=PINK, ha="left", color="FFFFFF", sz=10)

        h1, h2 = s2_title + 1, s2_title + 2
        ws.row_dimensions[h1].height = ws.row_dimensions[h2].height = 36

        for ci, hdr in [(2,"Sl\nNo"),(3,"Date"),(4,"Voucher\nNo."),(5,"Voucher\nAmount (Nu.)"),(6,"Details"),(REM_COL,"Remarks")]:
            ws.merge_cells(start_row=h1, start_column=ci, end_row=h2, end_column=ci)
            sc(ws.cell(h1, ci, hdr), bold=True, f=BHDR, ha="center", color="FFFFFF", sz=10)

        risk_groups = {}
        for col_i, (rt, _) in enumerate(rac, DYN_START):
            risk_groups.setdefault(rt, []).append(col_i)

        for rt, cols in risk_groups.items():
            col_s, col_e = cols[0], cols[-1]
            if col_s != col_e:
                ws.merge_cells(start_row=h1, start_column=col_s, end_row=h1, end_column=col_e)
            c = ws.cell(h1, col_s, rt)
            c.font = Font(bold=True, size=10, name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = YELLOW
            for ci2 in cols:
                ws.cell(h1, ci2).border = tb()

        for col_i, (_, asrt) in enumerate(rac, DYN_START):
            c = ws.cell(h2, col_i, asrt)
            c.font = Font(bold=True, size=10, name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = YELLOW; c.border = tb()

        dr_start, dr_end = h2 + 1, h2 + 20
        ynv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(ynv)

        for i, r in enumerate(range(dr_start, dr_end + 1)):
            rf = STRIPE if i % 2 else WHITE
            ws.row_dimensions[r].height = 18
            for c in range(2, LAST_COL + 1):
                cell = ws.cell(r, c)
                cell.border = tb(); cell.fill = rf
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
            ws.cell(r, 2, i + 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r, 2).font = Font(size=10, name="Arial"); ws.cell(r, 2).fill = rf
            for c in range(DYN_START, REM_COL):
                ynv.add(ws.cell(r, c))

        outer_med(ws, s2_title, dr_end, 2, LAST_COL)

        conc = dr_end + 3
        ws.row_dimensions[dr_end + 2].height = 6
        ws.row_dimensions[conc].height = 60
        c = ws.cell(conc, 2, "Overall Conclusion:")
        c.font = Font(bold=True, size=10, name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = GREEN; c.border = tb()
        ws.merge_cells(start_row=conc, start_column=3, end_row=conc, end_column=LAST_COL)
        ws.cell(conc, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(conc, 3).fill = WHITE
        for ci in range(3, LAST_COL + 1):
            ws.cell(conc, ci).border = tb()
        outer_med(ws, conc, conc, 2, LAST_COL)

    out = BytesIO()
    wb.save(out); out.seek(0)
    return out

def steps_bar(current: int):
    labels, icons = ["Upload", "Automate", "Download"], ["☁", "⚙", "⬇"]
    html = '<div class="steps-bar">'
    for i, (lbl, icon) in enumerate(zip(labels, icons), 1):
        nc  = "done" if i < current else ("active" if i == current else "")
        dot = "✓" if i < current else icon
        html += f'<div class="step-item"><div class="step-num {nc}">{dot}</div><span class="step-text {nc}">{lbl}</span></div>'
        if i < len(labels):
            html += f'<div class="step-connector {"done" if i < current else ""}"></div>'
    st.markdown(html + '</div>', unsafe_allow_html=True)

# ===================================================================
# MAIN TOOL PAGE
# ===================================================================
def tool_page():
    logo_path = "AWP Logo.png"

    # Hero badge
    st.markdown("""
    <div class="hero">
        <div class="badge"><div class="badge-dot"></div>Automation at Work</div>
    </div>
    """, unsafe_allow_html=True)

    # Logo
    if os.path.exists(logo_path):
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.image(logo_path, use_container_width=True)
    else:
        st.info("Logo not found. Place 'awp_5.2-removebg-preview.png' in the same directory.")

    # Hero heading
    st.markdown("""
    <div class="hero">
        <h1><span class="green">Smarter.</span><span class="green"> Faster.</span><span class="green"> Better.</span></h1>
        <p>Upload your AWP 4.6 Risk Response PDF and Let Automation Do The Rest.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Session state init ──────────────────────────────────────────
    for key, default in [("uploaded_pdf", None), ("excel_output", None), ("excel_ready", False)]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ── STEP 1 : Upload ────────────────────────────────────────────
    if st.session_state.uploaded_pdf is None:
        steps_bar(1)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            uploaded_pdf = st.file_uploader("⬆  Choose PDF File", type=["pdf"], label_visibility="collapsed")
            if uploaded_pdf is not None:
                st.session_state.uploaded_pdf = uploaded_pdf
                st.session_state.excel_ready  = False
                st.session_state.excel_output = None
                st.rerun()

    # ── STEP 2 : Extract ───────────────────────────────────────────
    elif not st.session_state.excel_ready:
        steps_bar(2)

        # PDF success card with inline Remove button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown(f"""
            <div class="upload-success">
                ✅ <strong>PDF uploaded successfully</strong><br>
                <span class="filename">{st.session_state.uploaded_pdf.name}</span>
            </div>
            """, unsafe_allow_html=True)
            st.markdown('<div class="remove-btn">', unsafe_allow_html=True)
            if st.button("🗑 Remove PDF", key="remove_pdf"):
                st.session_state.uploaded_pdf = None
                st.session_state.excel_ready  = False
                st.session_state.excel_output = None
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        default_name = re.sub(r'\.pdf$', '', st.session_state.uploaded_pdf.name, flags=re.IGNORECASE)
        st.session_state.file_name_input = st.text_input(
            "Output filename (no extension needed)",
            value=st.session_state.get("file_name_input", default_name or "FINAL_Audit_Workbook"),
            placeholder="e.g. BCTA_Audit_2024"
        )

        if st.button("⚙️  Extract to Excel"):
            with st.status("Processing your PDF…", expanded=True) as status:
                st.write("📖 Reading PDF tables…")
                st.session_state.excel_output = generate_excel(st.session_state.uploaded_pdf)
                st.write("🗂️ Building workbook structure…")
                st.write("🎨 Applying formatting…")
                status.update(label="✅ Done!", state="complete", expanded=False)
            st.session_state.excel_ready = True
            st.rerun()

    # ── STEP 3 : Download ──────────────────────────────────────────
    else:
        steps_bar(3)

        safe_name = re.sub(r'[\/*?:"<>|]', '_', st.session_state.get("file_name_input", "FINAL_Audit_Workbook")).strip() or "FINAL_Audit_Workbook"

        st.markdown('<div class="result-box"><div class="result-title">✅ Excel generated successfully</div>', unsafe_allow_html=True)
        st.download_button(
            label=f"⬇️  Download {safe_name}.xlsx",
            data=st.session_state.excel_output,
            file_name=f"{safe_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown('<div style="margin-top:0.75rem;"></div>', unsafe_allow_html=True)
        st.markdown('<div class="remove-btn">', unsafe_allow_html=True)
        if st.button("↩  Process Another PDF"):
            st.session_state.uploaded_pdf = None
            st.session_state.excel_ready  = False
            st.session_state.excel_output = None
            st.rerun()
        st.markdown('</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="app-footer">© Royal Audit Authority · Supreme Audit Institution of Bhutan</div>', unsafe_allow_html=True)

tool_page()
